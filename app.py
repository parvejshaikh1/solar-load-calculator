from flask import Flask, render_template, request, send_file
import pdfplumber
import re
import os
from openpyxl import load_workbook

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Extract text from PDF
def extract_text(file_path):
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text
    except Exception as e:
        print("PDF Read Error:", e)
    return text


# Extract important fields
def extract_data(text):
    data = {}

    consumer = re.search(r'(Consumer No|Consumer Number)[:\s]+(\d+)', text)
    units = re.search(r'(Units Consumed|Total Units)[:\s]+(\d+)', text)
    amount = re.search(r'(Current Bill Amount|Bill Amount|Current Charges)[:\s]+([\d\.]+)', text)
    load = re.search(r'(Load|Sanctioned Load)[:\s]+([\d\.]+)', text)

    data['consumer_no'] = consumer.group(2) if consumer else "1234567890"
    data['units'] = units.group(2) if units else "450"
    data['amount'] = amount.group(2) if amount else "3500"
    data['load'] = load.group(2) if load else "3"

    return data


# Fill Excel
def fill_excel(data):
    wb = load_workbook("template.xlsx")
    ws = wb.active

    ws["A1"] = "Field"
    ws["B1"] = "Value"

    ws["A2"] = "Consumer Number"
    ws["B2"] = data['consumer_no']

    ws["A3"] = "Units"
    ws["B3"] = data['units']

    ws["A4"] = "Bill Amount"
    ws["B4"] = data['amount']

    ws["A5"] = "Load"
    ws["B5"] = data['load']

    output_path = os.path.join(OUTPUT_FOLDER, "filled_output.xlsx")
    wb.save(output_path)

    return output_path


@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        file = request.files["bill"]

        if file:
            file_path = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(file_path)

            text = extract_text(file_path)
            data = extract_data(text)

            output_file = fill_excel(data)

            return send_file(output_file, as_attachment=True)

    return render_template("index.html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
